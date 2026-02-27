import io
from datetime import date
from django.contrib import admin
from django.utils.html import format_html
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from .models import (
    Level, ClassYear,
    Term, Subject, Student,TeacherProfile,
    Score, AcademicReport, StudentReportComment, MockReportComment
    # MidtermReport,
    # ProgressiveTestOneReport, ProgressiveTestTwoReport,
    # ProgressiveTestThreeReport
)

def _split_subjects(subjects_qs, break_after=7):
    names = [s.name for s in subjects_qs]
    if not names:
        return "None"
    groups = [names[i:i + break_after] for i in range(0, len(names), break_after)]
    return "\n".join(", ".join(g) for g in groups)


def _export_excel_response(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = ["#", "Student Full Name", "Student Class", "Student Subjects"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, student in enumerate(queryset, start=2):
        row_num = row_idx - 1
        class_display = str(student.class_year) if student.class_year else "N/A"
        subjects_display = _split_subjects(student.subjects.all())
        ws.cell(row=row_idx, column=1, value=row_num)
        ws.cell(row=row_idx, column=2, value=student.fullname)
        ws.cell(row=row_idx, column=3, value=class_display)
        subj_cell = ws.cell(row=row_idx, column=4, value=subjects_display)
        subj_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_idx].height = None

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 60

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"students_{date.today()}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _export_pdf_response(queryset):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'],
                                 fontSize=16, spellCheck=0,
                                 textColor=colors.HexColor("#1F4E79"),
                                 alignment=TA_CENTER, spaceAfter=4)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'],
                                    fontSize=10, alignment=TA_CENTER, spaceAfter=12)
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'],
                                fontSize=9, fontName='Helvetica', wordWrap='LTR')

    elements = []
    elements.append(Paragraph("Temple Christian International School", title_style))
    elements.append(Paragraph(f"Student List — Generated: {date.today().strftime('%B %d, %Y')}", subtitle_style))
    elements.append(Spacer(1, 0.4*cm))

    table_data = [["#", "Student Full Name", "Student Class", "Student Subjects"]]
    for row_num, student in enumerate(queryset, start=1):
        class_display = str(student.class_year) if student.class_year else "N/A"
        subjects_text = _split_subjects(student.subjects.all())
        subjects_para = Paragraph(subjects_text.replace("\n", "<br/>"), cell_style)
        table_data.append([row_num, student.fullname, class_display, subjects_para])

    col_widths = [1.2*cm, 7.0*cm, 5.0*cm, 13.5*cm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 10),
        ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME',   (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',   (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF3FB")]),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    filename = f"students_{date.today()}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def export_students_to_excel(modeladmin, request, queryset):
    qs = queryset.select_related('class_year', 'class_year__level') \
                 .prefetch_related('subjects') \
                 .order_by('fullname')
    return _export_excel_response(qs)

export_students_to_excel.short_description = "Export selected students to Excel"


def export_all_students_to_excel(modeladmin, request, queryset):
    qs = Student.objects.select_related('class_year', 'class_year__level') \
                        .prefetch_related('subjects') \
                        .order_by('fullname')
    return _export_excel_response(qs)

export_all_students_to_excel.short_description = "Export ALL students to Excel (alphabetical)"


def export_students_to_pdf(modeladmin, request, queryset):
    qs = queryset.select_related('class_year', 'class_year__level') \
                 .prefetch_related('subjects') \
                 .order_by('fullname')
    return _export_pdf_response(qs)

export_students_to_pdf.short_description = "Export selected students to PDF"


def export_all_students_to_pdf(modeladmin, request, queryset):
    qs = Student.objects.select_related('class_year', 'class_year__level') \
                        .prefetch_related('subjects') \
                        .order_by('fullname')
    return _export_pdf_response(qs)

export_all_students_to_pdf.short_description = "Export ALL students to PDF (alphabetical)"


# Register the Level model
class LevelAdmin(admin.ModelAdmin):
    list_display = ('name',)
    search_fields = ('name',)



# Register the ClassYear model
class ClassYearAdmin(admin.ModelAdmin):
    list_display = ('level', 'name')
    search_fields = ('name',)
    list_filter = ('level',)



# Register the Term model
class TermAdmin(admin.ModelAdmin):
    list_display = ('term_name', 'class_year')
    search_fields = ('term_name', 'class_year__name')
    list_filter = ('term_name',)



# Register the Subject model
class SubjectAdmin(admin.ModelAdmin):
    list_display = ('name', 'grading_system', 'get_class_years')
    search_fields = ('name',)
    list_filter = ('grading_system', 'class_year',)
    filter_horizontal = ('class_year',)

    def get_class_years(self, obj):
        return ", ".join([cy.name for cy in obj.class_year.all()])
    get_class_years.short_description = 'Class Years'



# Register the Student model
class StudentAdmin(admin.ModelAdmin):
    list_display = ('fullname', 'class_year')
    search_fields = ('fullname', 'class_year__name')
    list_filter = ('class_year',)
    filter_horizontal = ('subjects',)
    actions = [
        export_students_to_excel,
        export_all_students_to_excel,
        export_students_to_pdf,
        export_all_students_to_pdf,
    ]



# TeacherProfile Admin
class TeacherProfileAdmin(admin.ModelAdmin):
    list_display = ('user', 'get_subjects', 'get_class_years')
    search_fields = ('user__username', 'user__first_name', 'user__last_name')
    list_filter = ('subjects',)
    filter_horizontal = ('subjects',)
    
    def get_subjects(self, obj):
        return ", ".join([subject.name for subject in obj.subjects.all()])
    get_subjects.short_description = 'Subjects'

    def get_class_years(self, obj):
        class_years = set()
        for subject in obj.subjects.all():
            for class_year in subject.class_year.all():
                class_years.add(class_year.name)  # Assuming 'name' is a field of ClassYear
        return ", ".join(class_years)
    get_class_years.short_description = 'Class Years'




# Register the Score model
class ScoreAdmin(admin.ModelAdmin):
    list_display = ('student', 'subject', 'get_grading_system', 'term', 'class_work_score', 'progressive_test_1_score',
                    'progressive_test_2_score', 'midterm_score','mock_score', 'exam_score',
                    'continuous_assessment', 'total_score', 'grade')
    search_fields = ('student__fullname', 'subject__name', 'term__term_name')
    list_filter = ('term', 'subject', 'subject__grading_system', 'created_by')
    readonly_fields = ('continuous_assessment', 'total_score', 'get_grading_system')

    def get_grading_system(self, obj):
        return obj.subject.get_grading_system_display() if obj.subject else 'N/A'
    get_grading_system.short_description = 'Grading System'



# Register the StudentReportComment model
class StudentReportCommentAdmin(admin.ModelAdmin):
    list_display = ('student', 'class_year', 'term', 'has_academic_comment', 'has_behavioral_comment', 'created_by', 'updated_at')
    search_fields = ('student__fullname', 'class_year__name', 'term__term_name')
    list_filter = ('class_year', 'term', 'created_by')
    readonly_fields = ('created_at', 'updated_at')

    def has_academic_comment(self, obj):
        return bool(obj.academic_comment)
    has_academic_comment.boolean = True
    has_academic_comment.short_description = 'Academic'

    def has_behavioral_comment(self, obj):
        return bool(obj.behavioral_comment)
    has_behavioral_comment.boolean = True
    has_behavioral_comment.short_description = 'Behavioral'


# Register the MockReportComment model
class MockReportCommentAdmin(admin.ModelAdmin):
    list_display = ('student', 'class_year', 'term', 'has_academic_comment', 'has_behavioral_comment', 'created_by', 'updated_at')
    search_fields = ('student__fullname', 'class_year__name', 'term__term_name')
    list_filter = ('class_year', 'term', 'created_by')
    readonly_fields = ('created_at', 'updated_at')

    def has_academic_comment(self, obj):
        return bool(obj.academic_comment)
    has_academic_comment.boolean = True
    has_academic_comment.short_description = 'Academic'

    def has_behavioral_comment(self, obj):
        return bool(obj.behavioral_comment)
    has_behavioral_comment.boolean = True
    has_behavioral_comment.short_description = 'Behavioral'


# Register the AcademicReport model
class AcademicReportAdmin(admin.ModelAdmin):
    list_display = ('student', 'term', 'student_gpa')
    search_fields = ('student__fullname', 'term__term_name')
    list_filter = ('term',)





# Registering the models with custom admin classes
admin.site.register(Level, LevelAdmin)
admin.site.register(ClassYear, ClassYearAdmin)
admin.site.register(Term, TermAdmin)
admin.site.register(Student, StudentAdmin)
admin.site.register(TeacherProfile, TeacherProfileAdmin)
admin.site.register(Subject, SubjectAdmin)
admin.site.register(Score, ScoreAdmin)
admin.site.register(StudentReportComment, StudentReportCommentAdmin)
admin.site.register(MockReportComment, MockReportCommentAdmin)
# admin.site.register(MidtermReport, MidtermReportAdmin)
# admin.site.register(ProgressiveTestOneReport, ProgressiveTestOneReportAdmin)
# admin.site.register(ProgressiveTestTwoReport, ProgressiveTestTwoReportAdmin)
# admin.site.register(ProgressiveTestThreeReport, ProgressiveTestThreeReportAdmin)
admin.site.register(AcademicReport, AcademicReportAdmin)